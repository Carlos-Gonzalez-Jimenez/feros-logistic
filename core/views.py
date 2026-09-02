import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pandas as pd
from django.core.cache import cache
from django.db import models as output_field
from django.db import transaction
from django.db.models import Exists, ProtectedError, QuerySet, Value
from django.db.models import Prefetch
from django.db.models import (
    Subquery,
    OuterRef,
    F,
    ExpressionWrapper,
    DecimalField,
    Q,
    Sum,
    Avg,
    Count,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from django.utils.text import slugify
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.generics import GenericAPIView
from rest_framework.generics import (
    RetrieveUpdateAPIView,
    ListAPIView,
)
from rest_framework.permissions import (
    AllowAny,
    IsAuthenticatedOrReadOnly,
)
from rest_framework.response import Response

from core import models, serializers, filters
from core.exceptions import (
    ProtectedInstanceException,
    OrderUpdateException,
    StatusUnAvalaibleException,
    RollBackUnAvailableException,
    InvalidParameterException,
)
from core.tasks import (
    NomenclatorCacheManager,
    OrderInvoice,
    DeliveryDrive,
    ProductReport,
)
from dashboard.exceptions import StartDateCanNotBeAfterEnddateException
from delivery.models import ShippingRate, OrderShipping
from delivery.serializers import ShippingRateSerializer
from logistic_backend.settings import MEDIA_URL
from payments.exceptions import PaymentNotCompletedException
from payments.models import Payment, TransactionLog, Wallet, WalletOperationalLog
from payments.services import PaymentFactory
from user.models import Fee
from user.tasks import send_mail
from .builders import CreateOrderBuilder
from .generics import MultiplePermissionsView
from .odoo import sync_order_with_odoo_task
from .permissions import (
    CustomPermissionFactory,
    ReadOnlyPermission,
    ClientPermission,
    StaffPermission,
)
from .services import (
    NotificationService,
    get_state_handler
)


class ProtectedResourceViewSet(viewsets.ModelViewSet):
    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            instance = self.get_object()
            try:
                self.perform_destroy(instance)
                return Response(status=status.HTTP_204_NO_CONTENT)

            except ProtectedError as exception:
                raise ProtectedInstanceException() from exception


class BatchViewSet(viewsets.ModelViewSet):
    """
    Batch model\n
    GET: Shows all Batch created.\n
    POST: Adds a new Batch.\n
    GET{id}: Retrieves a specific Batch determined by id.\n
    PUT{id}: Modifies all fields of a specific Batch determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Batch determined by id.\n
    DELETE{id}: Deletes a specific Batch determined by id.\n
    """

    permission_classes = [CustomPermissionFactory(["core.manage_batch_entry"])]
    queryset = models.Batch.objects.annotate(
        break_event_point=Coalesce(
            Sum(F("items__quantity") * F("items__cost_price")),
            Value(0, output_field=output_field.DecimalField()),
        ),
    ).all()
    serializer_class = serializers.BatchSerializer
    filterset_class = filters.BatchFilter
    search_fields = ["batch_identificator", "notes"]

    @action(
        methods=["get"],
        detail=True,
        url_path="items",
        permission_classes=[CustomPermissionFactory(["core.manage_batch_entry"])],
    )
    def batch_items_all(self, request, pk=None):
        items = models.BatchItem.objects.filter(batch_id=pk)
        return Response(
            serializers.BatchItemSerializer(items, many=True, context=self.get_serializer_context()).data,
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="template")
    def download_template(self, request):
        """
        Descarga una plantilla Excel para cargar BatchItems
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Productos del lote"

        headers = [
            "Código / No. de Parte",
            "Nombre del Producto",
            "Cantidad",
            "Precio Costo USD",
        ]
        column_widths = [25, 40, 20, 20]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            column_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = column_widths[col_num - 1]

        response = BytesIO()
        workbook.save(response)

        return HttpResponse(
            response.getvalue(),
            content_type="application/ms-excel",
            headers={"Content-Disposition": "attachment; filename=productos_lote.xlsx"},
        )

    @action(
        methods=["post"],
        detail=True,
        url_path="import-batch-items",
        permission_classes=[CustomPermissionFactory(["core.manage_batch_entry"])],
    )
    def import_batch_items(self, request, pk):
        batch = self.get_object()
        serializer = serializers.BatchImportSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            with transaction.atomic():
                for batch_item_data in serializer.validated_data["batch_items_file"]:
                    item_product = models.Product.objects.filter(
                        part_code=batch_item_data["part_code"]
                    ).first()
                    item_name = batch_item_data["name"]
                    item_cost_price = batch_item_data["cost_price"]
                    item_part_code = batch_item_data["part_code"]
                    item_quantity = batch_item_data["quantity"]

                    batch_item = models.BatchItem.objects.filter(
                        Q(part_code=item_part_code) | Q(product=item_product)
                    ).first()
                    if batch_item:
                        batch_item.quantity = F("quantity") + item_quantity
                        batch_item.save(update_fields=["quantity", "cost_price"])
                    else:
                        models.BatchItem.objects.create(
                            sale_price=0,
                            batch=batch,
                            product=item_product,
                            quantity=item_quantity,
                            name=getattr(item_product, "name", item_name),
                            cost_price=batch.exchange_rate * item_cost_price,
                            part_code=getattr(
                                item_product, "part_code", item_part_code
                            ),
                        )

                all_items = models.BatchItem.objects.filter(batch_id=pk)
                return Response(
                    serializers.BatchItemSerializer(
                        all_items, many=True, context=self.get_serializer_context()
                    ).data,
                    status=status.HTTP_200_OK,
                )

    def _update_inventory(self, items: QuerySet[models.BatchItem]) -> None:
        products_to_update = []
        for item in items:
            product = item.product

            current_quantity = product.quantity
            current_cost = product.cost_price
            current_price = product.unit_price
            new_quantity = item.quantity
            new_cost = item.cost_price
            new_price = item.sale_price

            total_quantity = current_quantity + new_quantity

            if total_quantity > 0:
                weighted_cost = (
                                        (current_quantity * current_cost) + (new_quantity * new_cost)
                                ) / total_quantity
                weighted_price = (
                                         (current_quantity * current_price) + (new_quantity * new_price)
                                 ) / total_quantity
            else:
                weighted_cost = new_cost
                weighted_price = new_price

            product.quantity = total_quantity
            product.cost_price = weighted_cost
            product.unit_price = weighted_price
            products_to_update.append(product)

        models.Product.objects.bulk_update(products_to_update, ["quantity", "cost_price", "unit_price"])

    @action(
        methods=["post"],
        detail=True,
        url_path="update-inventory",
        permission_classes=[CustomPermissionFactory(["core.manage_batch_entry"])],
    )
    def update_inventory(self, request, pk=None):
        with transaction.atomic():
            items = models.BatchItem.objects.filter(batch_id=pk)
            self._update_inventory(items)
            models.Batch.objects.filter(pk=pk, processed=False).update(processed=True, processed_at=now())
        return Response(status=status.HTTP_200_OK)


class BatchItemViewSet(viewsets.ModelViewSet):
    """
    Batch Item model\n
    GET: Shows all Batch Item created.\n
    POST: Adds a new Batch Item.\n
    GET{id}: Retrieves a specific Batch Item determined by id.\n
    PUT{id}: Modifies all fields of a specific Batch Item determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Batch Item determined by id.\n
    DELETE{id}: Deletes a specific Batch Item determined by id.\n
    """

    permission_classes = [CustomPermissionFactory(["core.manage_batch_entry"])]
    queryset = models.BatchItem.objects.all()
    serializer_class = serializers.BatchItemSerializer

    def all_batchs_response(self, batch_id):
        batch_items = models.BatchItem.objects.filter(batch_id=batch_id)

        return Response(
            serializers.BatchItemSerializer(batch_items, many=True, context=self.get_serializer_context()).data,
            status=status.HTTP_200_OK
        )

    def create(self, request):
        with transaction.atomic():
            serializer = self.serializer_class(data=request.data, context=self.get_serializer_context())
            serializer.is_valid(raise_exception=True)
            item = serializer.save()
            self._merge_duplicate_batchitems(item)
            return self.all_batchs_response(item.batch_id)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            partial = kwargs.pop("partial", False)
            instance = self.get_object()

            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            item = serializer.save()
            self._merge_duplicate_batchitems(instance)
            return self.all_batchs_response(item.batch_id)

    def destroy(self, request, pk):
        with transaction.atomic():
            instance = self.get_object()
            batch_id = instance.batch_id
            self.perform_destroy(instance)
            return self.all_batchs_response(batch_id)

    def _merge_duplicate_batchitems(self, batch_item):
        product_id = batch_item.product_id
        if product_id:
            rows = (
                models.BatchItem.objects.filter(batch_id=batch_item.batch_id, product_id=product_id)
                .exclude(pk=batch_item.pk)
                .update(quantity=F("quantity") + batch_item.quantity)
            )
            if rows > 0:
                batch_item.delete()


class CurrencyViewSet(ProtectedResourceViewSet):
    """
    Currency model\n
    GET: Shows all Currencies created.\n
    POST: Adds a new Currency.\n
    GET{id}: Retrieves a specific Currency determined by id.\n
    PUT{id}: Modifies all fields of a specific Currency determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Currency determined by id.\n
    DELETE{id}: Deletes a specific Currency determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_currencies"]),
    ]
    queryset = models.Currency.objects.all()
    serializer_class = serializers.CurrencySerializer
    search_fields = ["name", "initials"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "currency", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "currency",
                "list",
                request.user,
                timeout=60 * 60 * 24,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "currency", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "currency",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24 * 7,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("currency")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("currency")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("currency")
        response = super().perform_destroy(instance)
        return response


class CreditTypeViewSet(ProtectedResourceViewSet):
    """
    CreditType model\n
    GET: Shows all Credit Types created.\n
    POST: Adds a new Credit Type.\n
    GET{id}: Retrieves a specific Credit Type determined by id.\n
    PUT{id}: Modifies all fields of a specific Credit Type determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Credit Type determined by id.\n
    DELETE{id}: Deletes a specific Credit Type determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_credittypes"]),
    ]
    queryset = models.CreditType.objects.all()
    serializer_class = serializers.CreditTypeSerializer
    search_fields = ["name"]

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "credittype", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "credittype",
                "list",
                request.user,
                timeout=60 * 60 * 24,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "credittype", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "credittype",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24 * 7,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("credittype")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("credittype")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("credittype")
        response = super().perform_destroy(instance)
        return response


class CountryViewSet(ProtectedResourceViewSet):
    """
    Country model\n
    GET: Shows all Countries created.\n
    POST: Adds a new Country.\n
    GET{id}: Retrieves a specific Country determined by id.\n
    PUT{id}: Modifies all fields of a specific Country determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Country determined by id.\n
    DELETE{id}: Deletes a specific Country determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_country"]),
    ]
    queryset = models.Country.objects.all()
    serializer_class = serializers.CountrySerializer
    search_fields = ["name", "code_alpha3"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "country", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "country",
                "list",
                request.user,
                timeout=60 * 60 * 24 * 7,  # 7 días
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "country", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "country",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24 * 30,  # 30 días
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("country")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("country")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("country")
        response = super().perform_destroy(instance)
        return response


class ProviderViewSet(ProtectedResourceViewSet):
    """
    Provider model\n
    GET: Shows all Providers created.\n
    POST: Adds a new Provider.\n
    GET{id}: Retrieves a specific Provider determined by id.\n
    PUT{id}: Modifies all fields of a specific Provider determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Provider determined by id.\n
    DELETE{id}: Deletes a specific Provider determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_provider"]),
    ]
    queryset = models.Provider.objects.all()
    serializer_class = serializers.ProviderSerializer
    search_fields = ["name"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "provider", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "provider",
                "list",
                request.user,
                timeout=60 * 60 * 12,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "provider", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "provider",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("provider")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("provider")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("provider")
        response = super().perform_destroy(instance)
        return response


class NotificationTypeViewSet(ProtectedResourceViewSet):
    """
    NotificationType model\n
    GET: Shows all Notification Types created.\n
    POST: Adds a new Notification Type.\n
    GET{id}: Retrieves a specific Notification Type determined by id.\n
    PUT{id}: Modifies all fields of a specific Notification Type determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Notification Type determined by id.\n
    DELETE{id}: Deletes a specific Notification Type determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission,
    ]
    queryset = models.NotificationType.objects.all()
    serializer_class = serializers.NotificationTypeSerializer


class NotificationViewSet(ProtectedResourceViewSet):
    """
    Notification model\n
    GET: Shows all Notifications created.\n
    POST: Adds a new Notification.\n
    GET{id}: Retrieves a specific Notification determined by id.\n
    PUT{id}: Modifies all fields of a specific Notification determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Notification determined by id.\n
    DELETE{id}: Deletes a specific Notification determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_notification"])
    ]
    queryset = models.Notification.objects.all()
    serializer_class = serializers.NotificationSerializer
    search_fields = ["title", "message"]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context


class NotificationUserViewSet(viewsets.ModelViewSet):
    queryset = models.NotificationUser.objects.all()
    serializer_class = serializers.NotificationUserSerializer
    filterset_class = filters.NotificationUserFilter

    @action(
        methods=["post"],
        detail=True,
        url_path="mark-as-read",
        permission_classes=[AllowAny],
    )
    def mark_notification_as_read(self, request, pk=None):
        with transaction.atomic():
            notification_user = self.get_object()
            if notification_user.user_id == request.user.id:
                notification_user.read = True
                notification_user.save()
                return Response(status=status.HTTP_200_OK)
            return Response(status=status.HTTP_404_NOT_FOUND)

    @action(
        methods=["post"],
        detail=False,
        url_path="mark-all-as-read",
        permission_classes=[AllowAny],
    )
    def mark_all_notifications_as_read(self, request, pk=None):
        with transaction.atomic():
            models.NotificationUser.objects.filter(
                user=request.user, read=False
            ).update(read=True)
            return Response(status=status.HTTP_200_OK)

    @action(
        methods=["get"], detail=False, url_path="status", permission_classes=[AllowAny]
    )
    def user_notifications_status(self, request, pk=None):
        with transaction.atomic():
            notifications_user = models.NotificationUser.objects.filter(
                user=request.user
            )
            data = {"unread": None, "total": notifications_user.count()}
            data["unread"] = notifications_user.filter(read=False).count()
            return Response(data, status=status.HTTP_200_OK)

    @action(
        methods=["get"],
        detail=False,
        url_path="received",
        permission_classes=[AllowAny],
    )
    def user_notifications_received(self, request, pk=None):
        with transaction.atomic():
            received = models.NotificationUser.objects.filter(user=request.user)
            filter_class = filters.NotificationUserFilter(
                request.query_params, queryset=received
            )
            received = filter_class.qs
            paginated = self.paginate_queryset(received)
            serializer = self.get_serializer(paginated, many=True)
            return self.get_paginated_response(serializer.data)


class BrandViewSet(ProtectedResourceViewSet):
    """
    Brand model\n
    GET: Shows all Brands created.\n
    POST: Adds a new Brand.\n
    GET{id}: Retrieves a specific Brand determined by id.\n
    PUT{id}: Modifies all fields of a specific Brand determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Brand determined by id.\n
    DELETE{id}: Deletes a specific Brand determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_brand"]),
    ]
    queryset = models.Brand.objects.all()
    serializer_class = serializers.BrandSerializer
    search_fields = ["name", "description"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    # def list(self, request, *args, **kwargs):
    #     page = request.query_params.get("page")
    #     page_size = request.query_params.get("page_size")
    #     search_term = request.query_params.get("search", "")

    #     if search_term and search_term.strip():
    #         return super().list(request, *args, **kwargs)

    #     cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

    #     cached_data = NomenclatorCacheManager.get_cached_data(
    #         "brand", "list", request.user, **cache_kwargs
    #     )

    #     if cached_data is not None:
    #         return Response(cached_data)

    #     response = super().list(request, *args, **kwargs)

    #     if response.status_code == 200:
    #         NomenclatorCacheManager.set_cached_data(
    #             response.data,
    #             "brand",
    #             "list",
    #             request.user,
    #             timeout=60 * 60 * 12,
    #             **cache_kwargs,
    #         )

    #     return response

    # def retrieve(self, request, *args, **kwargs):
    #     cached_data = NomenclatorCacheManager.get_cached_data(
    #         "brand", "retrieve", request.user, kwargs.get("pk")
    #     )

    #     if cached_data is not None:
    #         return Response(cached_data)

    #     response = super().retrieve(request, *args, **kwargs)

    #     if response.status_code == 200:
    #         NomenclatorCacheManager.set_cached_data(
    #             response.data,
    #             "brand",
    #             "retrieve",
    #             request.user,
    #             pk=kwargs.get("pk"),
    #             timeout=60 * 60 * 24,
    #         )

    #     return response

    # def perform_create(self, serializer):
    #     NomenclatorCacheManager.invalidate_model_cache("brand")
    #     response = super().perform_create(serializer)
    #     return response

    # def perform_update(self, serializer):
    #     NomenclatorCacheManager.invalidate_model_cache("brand")
    #     response = super().perform_update(serializer)
    #     return response

    # def perform_destroy(self, instance):
    #     NomenclatorCacheManager.invalidate_model_cache("brand")
    #     response = super().perform_destroy(instance)
    #     return response


class CategoryViewSet(ProtectedResourceViewSet):
    """
    Category model\n
    GET: Shows all Categories created.\n
    POST: Adds a new Category.\n
    GET{id}: Retrieves a specific Category determined by id.\n
    PUT{id}: Modifies all fields of a specific Category determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Category determined by id.\n
    DELETE{id}: Deletes a specific Category determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_category"])
    ]
    queryset = models.Category.objects.all()
    serializer_class = serializers.CategorySerializer
    search_fields = ["name"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    @action(
        detail=True,
        methods=["get"],
        url_path=r"subcategories",
        permission_classes=[AllowAny],
    )
    def sub_categories(self, request, pk=None):
        with transaction.atomic():
            category = self.get_object()
            sub_categories = models.Category.objects.filter(
                parent=category.id, active=True
            )
            return Response(
                serializers.CategorySerializer(
                    sub_categories, many=True, context={"request": request}
                ).data,
                status=status.HTTP_200_OK,
            )


class MeasurementUnitViewSet(ProtectedResourceViewSet):
    """
    Measurement Unit model\n
    GET: Shows all Measurement Units created.\n
    POST: Adds a new Measurement Unit.\n
    GET{id}: Retrieves a specific Measurement Unit determined by id.\n
    PUT{id}: Modifies all fields of a specific Measurement Unit determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Measurement Unit determined by id.\n
    DELETE{id}: Deletes a specific Measurement Unit determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_measurement_unit"])
    ]
    queryset = models.Measurement_Unit.objects.all()
    serializer_class = serializers.MeasurementUnitSerializer
    search_fields = ["name", "abbreviation"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "measurement_unit", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "measurement_unit",
                "list",
                request.user,
                timeout=60 * 60 * 12,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "measurement_unit", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "measurement_unit",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("measurement_unit")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("measurement_unit")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("measurement_unit")
        response = super().perform_destroy(instance)
        return response


class OrderStatusViewSet(ProtectedResourceViewSet):
    """
    Order Status model\n
    GET: Shows all order's statuses created.\n
    POST: Adds a new order's status.\n
    GET{id}: Retrieves a specific order's status determined by id.\n
    PUT{id}: Modifies all fields of a specific order's status determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific order's status determined by id.\n
    DELETE{id}: Deletes a specific order's status determined by id.\n
    """

    queryset = models.OrderStatus.objects.all()
    serializer_class = serializers.OrderStatusSerializer
    filterset_class = filters.StatusFilter
    search_fields = ["name"]


class SpecificationsViewSet(ProtectedResourceViewSet):
    """
    Specification model\n
    GET: Shows all specifications created.\n
    POST: Adds a new specification.\n
    GET{id}: Retrieves a specific specification determined by id.\n
    PUT{id}: Modifies all fields of a specific specification determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific specification determined by id.\n
    DELETE{id}: Deletes a specific specification determined by id.\n
    """

    queryset = models.Specifications.objects.all()
    serializer_class = serializers.SpecificationsSerializer
    search_fields = ["name"]

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(active=True)

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "specifications", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "specifications",
                "list",
                request.user,
                timeout=60 * 60 * 12,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "specifications", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "specifications",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("specifications")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("specifications")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("specifications")
        response = super().perform_destroy(instance)
        return response


class VehicleTypeViewSet(ProtectedResourceViewSet):
    """
    Vehicle Type model\n
    GET: Shows all vehicle's types created.\n
    POST: Adds a new vehicle's type.\n
    GET{id}: Retrieves a specific vehicle's type determined by id.\n
    PUT{id}: Modifies all fields of a specific vehicle's type determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific vehicle's type determined by id.\n
    DELETE{id}: Deletes a specific vehicle's type determined by id.\n
    """

    permission_classes = [CustomPermissionFactory(["core.manage_vehicles"])]
    queryset = models.VehicleType.objects.all()
    serializer_class = serializers.VehicleTypeSerializer


class VehicleViewSet(ProtectedResourceViewSet):
    """
    Vehicle model\n
    GET: Shows all vehicles created.\n
    POST: Adds a new vehicle.\n
    GET{id}: Retrieves a specific vehicle determined by id.\n
    PUT{id}: Modifies all fields of a specific vehicle determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific vehicle determined by id.\n
    DELETE{id}: Deletes a specific vehicle determined by id.\n
    """

    permission_classes = [CustomPermissionFactory(["core.manage_vehicles"])]
    queryset = models.Vehicle.objects.all()
    serializer_class = serializers.VehicleSerializer

    @action(
        detail=False,
        methods=["get"],
        url_path=r"tracking",
        permission_classes=[CustomPermissionFactory(["core.manage_vehicles"])],
    )
    def tracking(self, request):
        vehicles = models.Vehicle.objects.filter(active=True, locations__isnull=False).distinct()
        locations = []
        for vehicle in vehicles:
            try:
                location = vehicle.locations.latest('created_at')
                if location:
                    locations.append(location)
            except models.VehicleLocation.DoesNotExist:
                pass
        return Response(serializers.VehicleLocationSerializer(instance=locations, many=True).data)


class SpecificationDetailsViewSet(ProtectedResourceViewSet):
    """
    Specification Detail model\n
    GET: Shows all specification details created.\n
    POST: Adds a new specification detail.\n
    GET{id}: Retrieves a specific specification detail determined by id.\n
    PUT{id}: Modifies all fields of a specific specification detail determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific specification detail determined by id.\n
    DELETE{id}: Deletes a specific specification detail determined by id.\n
    """

    queryset = models.SpecificationDetails.objects.all()
    serializer_class = serializers.SpecificationDetailsSerializer


def check_category_existence(product_category: str):
    """check for a product category existence

    Args:
        category (str): categroy name

    Returns:
        Category: an existence category
    """

    first_point_position = product_category.find(".")
    parent_category_name = None
    category_name = None
    if first_point_position != -1:
        parent_category_name = product_category[0:first_point_position]
        category_name = product_category[first_point_position + 1:]
    else:
        category_name = product_category
    parent_category = None
    if parent_category_name is not None:
        parent_categories = models.Category.objects.filter(name=parent_category_name)
        if parent_categories.exists():
            parent_category = parent_categories.first()
        else:
            parent_category = models.Category.objects.create(
                name=parent_category_name,
                parent=None,
            )
    category = models.Category.objects.filter(
        name=category_name, parent=parent_category
    )
    if category.exists():
        category = category.first()
    else:
        category = models.Category(
            name=category_name,
            parent=parent_category,
        )
        category.save()
    return category


def check_m_unit_existence(m_unit: str):
    """check for a measurement unit existence

    Args:
        m_unit (str): measurement unit abbreviation

    Returns:
        Measument_Unit: an existence measurement unit
    """
    measurement_unit = models.Measurement_Unit.objects.filter(abbreviation=m_unit)
    if measurement_unit.exists():
        measurement_unit = measurement_unit.first()
    else:
        measurement_unit = models.Measurement_Unit(name=m_unit, abbreviation=m_unit)
        measurement_unit.save()
    return measurement_unit


class CacheInvalidator:
    cache_keys = ["landing_response"]

    @classmethod
    def invalidate_landing_cache(cls):
        for key in cls.cache_keys:
            cache.delete(key)


class ProductViewSet(ProtectedResourceViewSet):
    """
    Product model\n
    GET: Shows all Products created.\n
    POST: Adds a new Product.\n
    GET{id}: Retrieves a specific Product determined by id.\n
    PUT{id}: Modifies all fields of a specific Product determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Product determined by id.\n
    DELETE{id}: Deletes a specific Product determined by id.\n
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_product"])
    ]
    queryset = models.Product.objects.all()
    filterset_class = filters.ProductFilter
    search_fields = ["code_sku", "name"]

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return serializers.ProductWriteSerializer
        if self.action == "list":
            return serializers.ProductReadMinimalSerializer
        return serializers.ProductReadSerializer

    def perform_update(self, serializer):
        with transaction.atomic():
            response = super().perform_update(serializer)
            CacheInvalidator.invalidate_landing_cache()
            return response

    @action(
        methods=["post"],
        detail=False,
        url_path="import",
        permission_classes=[CustomPermissionFactory(["core.manage_product"])],
    )
    def import_excel(self, request):
        with transaction.atomic():
            if "excel" in request.FILES:
                excel = pd.read_excel(request.FILES["excel"], skiprows=11).fillna("")
                models.Product.objects.all().update(active=False)
                new_products = 0
                for index, _ in excel.iterrows():
                    category = excel.at[index, excel.columns[2]].strip()
                    code_sku = excel.at[index, excel.columns[3]]
                    name = excel.at[index, excel.columns[4]].strip()
                    m_unit = excel.at[index, excel.columns[5]].strip().upper()
                    quantity_per_box = excel.at[index, excel.columns[6]]
                    quantity = excel.at[index, excel.columns[7]]
                    unit_price = excel.at[index, excel.columns[8]]

                    products = models.Product.objects.filter(code_sku=code_sku)
                    if products.exists():
                        product = products.first()
                        price_diference = unit_price - product.unit_price
                        product.daily_variation = price_diference
                        product.quantity = "{:.2f}".format(quantity)
                        product.unit_price = "{:.2f}".format(unit_price)
                        product.quantity_per_box = "{:.2f}".format(quantity_per_box)
                    else:
                        category = check_category_existence(category)
                        m_unit = check_m_unit_existence(m_unit)
                        product = models.Product(
                            code_sku=code_sku,
                            name=name,
                            slug=slugify(name),
                            quantity_per_box="{:.2f}".format(quantity_per_box),
                            quantity="{:.2f}".format(quantity),
                            unit_price="{:.2f}".format(unit_price),
                            measurement_unit=m_unit,
                            category=category,
                            minimal_stock=30,
                            description="",
                            net_weight=0.0,
                            gross_weight=0.0,
                        )
                        new_products += 1
                    if float(product.quantity) > 0:
                        product.active = True
                    product.save()
                users = models.User.objects.filter(is_staff=True)
                NotificationService.send_notification(
                    "Importación de productos",
                    "Ha sido importado el archivo de productos",
                    users,
                    "Informativo",
                )
                if new_products > 0:
                    users = models.User.objects.filter(groups__name="Administrador")
                    NotificationService.send_notification(
                        "Nuevos productos",
                        f"La última importación ha incorporado <b>{new_products}</b> nuevo(s) producto(s) al inventario. Por favor revíselos",
                        users,
                        "Informativo",
                    )
            return Response(status=status.HTTP_200_OK)

    def get_category_tree(self):
        categories = models.Category.objects.filter(active=True).prefetch_related(
            Prefetch(
                "products",
                queryset=models.Product.objects.filter(active=True).only(
                    "code_sku",
                    "name",
                    "unit_price",
                    "quantity",
                    "active",
                    "minimal_stock",
                    "category_id",
                ),
            )
        )
        category_dict = {}
        root_categories = []
        for cat in categories:
            category_dict[cat.id] = {
                "obj": cat,
                "children": [],
                "products": list(cat.products.all()),
                "parent_id": cat.parent_id,
            }
        for _, cat_data in category_dict.items():
            parent_id = cat_data["parent_id"]

            if parent_id and parent_id in category_dict:
                category_dict[parent_id]["children"].append(cat_data)
            else:
                root_categories.append(cat_data)
        return root_categories

    def generate_product_report(self):
        pdf = ProductReport()
        pdf.add_page()
        category_tree = self.get_category_tree()
        total_products = 0
        total_categories = 0
        total_stock_value = Decimal("0.00")
        total_wholesale_value = Decimal("0.00")

        def process_category(category_data, level=0):
            nonlocal total_products, total_categories, total_stock_value, total_wholesale_value

            category = category_data["obj"]
            products = category_data["products"]
            children = category_data["children"]

            total_categories += 1
            pdf.category_header(category.name, level)

            if products:
                pdf.set_font("Roboto", "B", 9)
                pdf.set_x(10 + level * 5)
                pdf.cell(20, 8, "Código")
                pdf.cell(70, 8, "Producto")
                pdf.cell(25, 8, "P.Unitario", 0, 0, "R")
                pdf.cell(25, 8, "P.Mayorista", 0, 0, "R")
                pdf.cell(25, 8, "Cant.Mayorista", 0, 0, "R")
                pdf.cell(25, 8, "Existencias", 0, 0, "R")

            for product in products:
                if pdf.get_y() > 300:
                    pdf.add_page()

                pdf.product_row(product, level)
                total_products += 1
                total_stock_value += product.quantity * product.unit_price
                total_wholesale_value += product.quantity * product.wholesale_price

            for child in children:
                process_category(child, level + 1)
                pdf.ln(10)

        for root_category in category_tree:
            process_category(root_category)
            pdf.ln(10)

        pdf.summary_section(
            total_products, total_categories, total_stock_value, total_wholesale_value
        )
        return pdf

    @action(
        methods=["get"],
        detail=False,
        url_path="export",
        permission_classes=[CustomPermissionFactory(["user.show_availables_products"])],
    )
    def export_products(self, request):
        """
        Returns a pdf file with the information related to all active products.\n
        """
        pdf = self.generate_product_report()

        response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
        return response

    @action(
        methods=["get"],
        detail=True,
        url_path="reviews",
        permission_classes=[AllowAny],
    )
    def get_reviews(self, request, pk=None):
        """
        Returns all reviews from a product.
        """
        product = self.get_object()

        reviews = models.Review.objects.filter(product_id=product.id).order_by(
            "-review_date"
        )

        paginated_reviews = self.paginate_queryset(reviews)
        serializer = serializers.ReviewSerializer(
            paginated_reviews, many=True, context={"request": request}
        )

        return self.get_paginated_response(serializer.data)


class ShopViewSet(ListAPIView):
    permission_classes = [AllowAny]
    queryset = models.Product.objects.all()
    serializer_class = serializers.ProductReadSerializer
    filterset_class = filters.ProductFilter


class ProductSlugView(ListAPIView):
    permission_classes = [AllowAny]
    serializer_class = serializers.ProductReadSerializer

    def get_queryset(self):
        return models.Product.objects.select_related(
            "category",
            "brand",
            "provider",
            "country",
            "measurement_unit",
        ).filter(active=True)

    def get(self, request, slug):
        try:
            product = self.get_queryset().get(slug=slug)
            serializer = self.get_serializer(product)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except models.Product.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)


class CategoriesWithProductsView(ListAPIView):
    """_summary_

    Args:
        ListAPIView (_type_): _description_

    Returns:
        _type_: _description_
    """

    permission_classes = [AllowAny]
    queryset = models.Category.objects.all()
    serializer_class = serializers.CategorySerializer

    def list(self, request):
        """_summary_

        Args:
            request (_type_): _description_

        Returns:
            _type_: _description_
        """
        is_active = request.GET.get("active", "false") == "true"
        queryset = models.Product.objects.all()
        if is_active:
            queryset = queryset.filter(active=True)

        categories = models.Category.objects.filter(
            Q(id__in=queryset.values_list("category_id", flat=True)) | Q(parent=None),
            active=True,
        )

        serializer = serializers.CategorySerializer(
            categories,
            many=True,
            context={"request": request},
        )
        return Response(serializer.data, status=status.HTTP_200_OK)


class CountriesWithProductsView(ListAPIView):
    """_summary_

    Args:
        APIView (_type_): _description_

    Returns:
        _type_: _description_
    """

    permission_classes = [AllowAny]
    queryset = models.Country.objects.all()
    serializer_class = serializers.CountrySerializer

    def list(self, request):
        """_summary_

        Args:
            request (_type_): _description_

        Returns:
            _type_: _description_
        """
        countries = models.Country.objects.filter(product__active=True).distinct()

        serializer = serializers.CountrySerializer(
            countries,
            many=True,
            context={"request": request},
        )
        return Response(serializer.data, status=status.HTTP_200_OK)


def cart_total_amount(cart_products: list, fee: Fee) -> float:
    return sum(
        cart_product.quantity
        * (
            cart_product.product.sell_wholesale_price(fee)
            if cart_product.product.has_wholesale_price
               and cart_product.quantity >= cart_product.product.wholesale_minimum
            else cart_product.product.sell_price(fee)
        )
        for cart_product in cart_products
    )


def cart_total_gross_weight(cart_products: list) -> float:
    return sum(
        cart_product.quantity * cart_product.product.gross_weight
        for cart_product in cart_products
    )


def cart_total_net_weight(cart_products: list) -> float:
    return sum(
        cart_product.quantity * cart_product.product.net_weight
        for cart_product in cart_products
    )


def cart_total_boxes(cart_products: list) -> float:
    return sum(cart_product.quantity for cart_product in cart_products)


class CartViewSet(viewsets.ModelViewSet):
    """_summary_

    Args:
        ModelViewSet (_type_): _description_
    """

    permission_classes = [ClientPermission]
    queryset = models.Cart.objects.all()
    serializer_class = serializers.CartSerializer

    def get_response(self, client, status=status.HTTP_200_OK):
        cart_products = (
            models.Cart.objects.select_related("product")
            .prefetch_related("product__category")
            .filter(client_id=client.id)
        )

        total_amount = cart_total_amount(cart_products, client.fee)
        total_gross_weight = cart_total_gross_weight(cart_products)
        total_net_weight = cart_total_net_weight(cart_products)
        total_boxes = cart_total_boxes(cart_products)

        return Response(
            {
                "total_amount": total_amount,
                "total_gross_weight": total_gross_weight,
                "total_net_weight": total_net_weight,
                "total_boxes": total_boxes,
                "products": serializers.CartSerializer(
                    cart_products, many=True, context=self.get_serializer_context()
                ).data,
            },
            status=status,
        )

    def list(self, request):
        user = self.request.user
        return self.get_response(user, status.HTTP_200_OK)

    def create(self, request):
        with transaction.atomic():
            serializer = self.serializer_class(data=request.data, context=self.get_serializer_context())
            serializer.is_valid(raise_exception=True)

            product = serializer.save()
            client = product.client

            return self.get_response(client, status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            partial = kwargs.pop("partial", False)
            instance = self.get_object()
            client = instance.client

            serializer = self.get_serializer(
                instance, data=request.data, partial=partial
            )
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

            return self.get_response(client, status.HTTP_200_OK)

    def destroy(self, request, pk):
        with transaction.atomic():
            instance = self.get_object()
            client = instance.client
            self.perform_destroy(instance)
            return self.get_response(client, status.HTTP_200_OK)

    @action(
        detail=False,
        methods=["get"],
        url_path=r"clear",
        permission_classes=[ClientPermission],
    )
    def delete_cart(self, request):
        client = request.user
        with transaction.atomic():
            models.Cart.objects.filter(client_id=client.id).delete()
        return self.get_response(client, status.HTTP_200_OK)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"checkout",
        permission_classes=[ClientPermission],
    )
    def checkout(self, request):
        with transaction.atomic():
            serializer = serializers.CreateOrderSerializer(data=request.data)
            if serializer.is_valid(raise_exception=True):
                try:
                    user = self.request.user

                    builder = CreateOrderBuilder(models.Cart, serializer.validated_data, client=self.request.user, seller=None)
                    builder.create_order()
                    builder.append_order_products()
                    builder.apply_discount()
                    builder.add_shipping()
                    # builder.add_credit()
                    builder.add_expiration_days()
                    builder.clear_cart()

                    builder.update_order_amounts()
                    builder.add_initial_state('Estado inicial [Solicitud de compra creada]')
                    response, _ = builder.create_payment()

                    order = builder.order

                    NotificationService.send_notification(
                        "Nueva solicitud de compra",
                        f"El cliente <b>{user.first_name} {user.last_name}</b> ha creado una solicitud de compra nueva <b>{order}</b>",
                        [user],
                        "Informativo",
                        ["IN_APP", "WHATSAPP"],
                    )

                    builder.sync_with_odoo()
                    return Response(response, status=status.HTTP_200_OK)

                except models.Cart.DoesNotExist:
                    return Response(status=status.HTTP_404_NOT_FOUND)


def create_excel_file(products):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nueva solicitud de compra"

    styles = {
        "header": {
            "font": Font(bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="366092", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "title": {
            "font": Font(bold=True, size=16),
            "alignment": Alignment(horizontal="center"),
        },
    }

    ws.merge_cells("A1:D1")
    ws["A1"] = "Nueva solicitud de compra"
    ws["A1"].font = styles["title"]["font"]
    ws["A1"].alignment = styles["title"]["alignment"]

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="left")

    headers = ["Código", "Producto", "Cantidad", "Importe"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        for attr, value in styles["header"].items():
            setattr(cell, attr, value)

    for row_idx, product in enumerate(products, 5):
        ws.cell(row=row_idx, column=1, value=product.product.code_sku)
        ws.cell(row=row_idx, column=2, value=product.product.name)
        ws.cell(row=row_idx, column=3, value=product.quantity)
        ws.cell(row=row_idx, column=4, value=product.quantity * product.price)

    ws.column_dimensions["B"].width = 50

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


class OrderViewSet(viewsets.ModelViewSet):
    """_summary_

    Args:
        viewsets (_type_): _description_
    """

    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_order"])
    ]
    queryset = models.Order.objects.filter(merge__isnull=True)
    serializer_class = serializers.OrderSerializer
    filterset_class = filters.OrderFilter
    search_fields = ["id", "client__first_name", "client__last_name"]

    def get_serializer_class(self):
        if self.action == "list":
            return serializers.OrderMinimalSerializer
        return serializers.OrderSerializer

    def perform_update(self, serializer):
        super().perform_update(serializer)
        sync_order_with_odoo_task(serializer.instance)

    @action(
        detail=False,
        methods=['get'],
        url_path=r'pending',
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
        filterset_class=filters.PendingOrderFilter,
    )
    def pending(self, request):
        queryset = (
            models.Order.objects.annotate(
                status=Subquery(
                    models.OrderTracking.objects.filter(order=OuterRef('pk')).order_by('-id').values(
                        'status__code_name')[:1]
                )
            )
            .filter(expiration_date__gte=now())
            .exclude(status__in=['completed', 'cancelled', 'returned'])
            .order_by('expiration_date')
        )
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=["get"],
        url_path="invoice",
        permission_classes=[
            IsAuthenticatedOrReadOnly,
        ],
    )
    def invoice(self, request, pk=None):
        order = self.get_object()
        pdf = OrderInvoice(order)
        response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
        return response

    @action(
        detail=True,
        methods=["get"],
        url_path="delivery-drive",
        permission_classes=[IsAuthenticatedOrReadOnly],
    )
    def delivery_drive(self, request, pk=None):
        order = self.get_object()
        pdf = DeliveryDrive(order)
        response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
        return response

    @action(
        detail=True,
        methods=["post"],
        url_path=r"send-to-invoice",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def send_to_invoice(self, request, pk=None):
        with transaction.atomic():
            order = self.get_object()
            products = models.OrderProducts.objects.filter(order_id=order.id)
            if products.exists():
                config_settings = models.Config.objects.get()
                billing_email = config_settings.billing_email
                if billing_email is not None and billing_email != "":
                    excel_file = create_excel_file(products)
                    context = {
                        "logo": request.build_absolute_uri(
                            f"{MEDIA_URL}{config_settings.logo_light}"
                        ),
                        "client_name": f"{products[0].order.client.first_name} {products[0].order.client.last_name}",
                        "client_email": products[0].order.client.email,
                        "contact_name": f"{products[0].order.client.first_name} {products[0].order.client.last_name}",
                        "phone_number": products[0].order.client.phone_number,
                        "seller_name": "",
                        "seller_phone_number": "",
                        "products": products,
                        "business_name": config_settings.business_name,
                        "frontend_url": config_settings.login_url,
                        "total_amount": sum(
                            product.quantity * product.price for product in products
                        ),
                    }
                    message = get_template("mailing/purchase_request.html").render(
                        context
                    )
                    email_list = billing_email.split(";")
                    send_mail(
                        email_list,
                        "Nueva solicitud de compra",
                        message,
                        excel_file,
                    )
                else:
                    raise InvalidParameterException()
            return Response(
                serializers.OrderSerializer(
                    order, context=self.get_serializer_context()
                ).data,
                status=status.HTTP_200_OK,
            )

    @action(
        detail=True,
        methods=["post"],
        url_path=r"add-products",
        permission_classes=[CustomPermissionFactory(["core.manage_orderproducts"])],
    )
    def add_products(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        with transaction.atomic():
            product_ids = request.data.get("product_ids", [])

            if not product_ids:
                return Response(status=status.HTTP_400_BAD_REQUEST)

            existing_products = models.OrderProducts.objects.filter(
                order_id=order.id, product_id__in=product_ids
            ).values_list("product_id", flat=True)

            new_product_ids = set(product_ids) - set(existing_products)

            if not new_product_ids:
                return Response(status=status.HTTP_200_OK)

            new_products = models.Product.objects.filter(id__in=new_product_ids)
            product_map = {product.id: product for product in new_products}

            order_products_to_create = [
                models.OrderProducts(
                    order_id=order.id,
                    product_id=product_id,
                    quantity=1,
                    price=product_map[product_id].unit_price,
                )
                for product_id in new_product_ids
                if product_id in product_map
            ]

            models.OrderProducts.objects.bulk_create(order_products_to_create)
            total_products_added = len(order_products_to_create)

            if total_products_added > 0:
                NotificationService.send_notification(
                    "Solicitud actualizada",
                    f"La solicitud de compra <b>{order.id}</b> ha sido actualizada. "
                    f"Se le agregaron {total_products_added} nuevo(s) producto(s) a la solicitud",
                    [user, order.client],
                    "Informativo",
                    ["IN_APP", "WHATSAPP"],
                )

            order.refresh_from_db()

            return Response(
                serializers.OrderSerializer(
                    order, context=self.get_serializer_context()
                ).data,
                status=status.HTTP_200_OK,
            )

    @action(
        detail=True,
        methods=["post"],
        url_path=r"merge",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def merge_orders(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        with transaction.atomic():
            serializer = serializers.MergeOrderSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            orders_merge_to = serializer.validated_data["order_list"]
            delivery_address = serializer.validated_data["delivery_address_id"]
            shipping_rate = serializer.validated_data["shipping_rate_id"]
            observations = serializer.validated_data["observations"]

            if orders_merge_to is not None:
                new_order = models.Order.objects.create(
                    client_id=order.client_id,
                    seller=order.seller,
                    percentual_fee=order.percentual_fee,
                    fixed_fee=order.fixed_fee,
                    delivery_address=delivery_address,
                    observations=observations,
                )

                initial_status = models.OrderStatus.objects.get(initial_status=True)
                models.OrderTracking.objects.create(
                    order_id=new_order.id,
                    status_id=initial_status.id,
                    observations=f"Estado inicial [Solicitud creada por la unión de otras solicitudes ({orders_merge_to})]",
                )
                orders_merge_to.append(order.id)
                for order_id in orders_merge_to:
                    order_products = models.OrderProducts.objects.filter(
                        order_id=order_id
                    )
                    if order_products.exists():
                        for order_product in order_products:
                            product_to_merge = models.OrderProducts.objects.filter(
                                order_id=new_order.id,
                                product_id=order_product.product_id,
                            )
                            if product_to_merge.exists():
                                product_to_merge = product_to_merge.first()
                                product_to_merge.quantity += order_product.quantity
                                product_to_merge.save()
                            else:
                                _ = models.OrderProducts.objects.create(
                                    quantity=order_product.quantity,
                                    price=order_product.price,
                                    order_id=new_order.id,
                                    product_id=order_product.product_id,
                                )
                    order_to_merge = models.Order.objects.filter(id=order_id)
                    if order_to_merge.exists():
                        order_to_merge = order_to_merge.first()
                        order_to_merge.merge = new_order
                        order_to_merge.save()
                NotificationService.send_notification(
                    "Unión de solicitudes",
                    f"Nueva solicitud de compra creada <b>{new_order}</b> a partir de la unión de varias solicitudes existentes",
                    [user, order.client],
                    "Informativo",
                    ["IN_APP", "WHATSAPP"],
                )
                return Response(
                    serializers.OrderSerializer(
                        new_order, context=self.get_serializer_context()
                    ).data,
                    status=status.HTTP_200_OK,
                )

    # def merge_orders(self, request, pk=None):
    #     order = self.get_object()
    #     user = self.request.user
    #     with transaction.atomic():
    #         orders_merge_to = request.data["order_list"]
    #         delivery_address = request.data["delivery_address"]
    #         observations = request.data["observations"]
    #
    #         if orders_merge_to is not None:
    #             new_order = models.Order.objects.create(
    #                 client_id=order.client_id,
    #                 seller=order.seller,
    #                 percentual_fee=order.percentual_fee,
    #                 fixed_fee=order.fixed_fee,
    #                 delivery_address=delivery_address,
    #                 observations=observations,
    #             )
    #             initial_status = models.OrderStatus.objects.get(initial_status=True)
    #             models.OrderTracking.objects.create(
    #                 order_id=new_order.id,
    #                 status_id=initial_status.id,
    #                 observations=f"Estado inicial [Solicitud creada por la unión de otras solicitudes ({orders_merge_to})]",
    #             )
    #             orders_merge_to.append(order.id)
    #             for order_id in orders_merge_to:
    #                 order_products = models.OrderProducts.objects.filter(
    #                     order_id=order_id
    #                 )
    #                 if order_products.exists():
    #                     for order_product in order_products:
    #                         product_to_merge = models.OrderProducts.objects.filter(
    #                             order_id=new_order.id,
    #                             product_id=order_product.product_id,
    #                         )
    #                         if product_to_merge.exists():
    #                             product_to_merge = product_to_merge.first()
    #                             product_to_merge.quantity += order_product.quantity
    #                             product_to_merge.save()
    #                         else:
    #                             _ = models.OrderProducts.objects.create(
    #                                 quantity=order_product.quantity,
    #                                 price=order_product.price,
    #                                 order_id=new_order.id,
    #                                 product_id=order_product.product_id,
    #                             )
    #                 order_to_merge = models.Order.objects.filter(id=order_id)
    #                 if order_to_merge.exists():
    #                     order_to_merge = order_to_merge.first()
    #                     order_to_merge.merge = new_order
    #                     order_to_merge.save()
    #             NotificationService.send_notification(
    #                 "Unión de solicitudes",
    #                 f"Nueva solicitud de compra creada <b>{new_order}</b> a partir de la unión de varias solicitudes existentes",
    #                 [user, order.client],
    #                 "Informativo",
    #                 ["IN_APP", "WHATSAPP"],
    #             )
    #             return Response(
    #                 serializers.OrderSerializer(
    #                     new_order, context=self.get_serializer_context()
    #                 ).data,
    #                 status=status.HTTP_200_OK,
    #             )

    @action(
        detail=True,
        methods=["post"],
        url_path=r"roll-back",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def roll_back_order(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        with transaction.atomic():
            merge_orders = models.Order.objects.filter(merge_id=order.id)
            if merge_orders.exists():
                merge_orders.update(merge=None)
                NotificationService.send_notification(
                    "Restablecer solicitudes",
                    f"Solicitud de compra <b>{order.id}</b> restablecida. Se eliminó la solicitud y se restablecieron las solicitudes de compra originales unidas previamente",
                    [user, order.client],
                    "Informativo",
                    ["IN_APP", "WHATSAPP"],
                )
                order.delete()
                return Response(status=status.HTTP_200_OK)
            raise RollBackUnAvailableException()

    @action(
        detail=True,
        methods=["post"],
        url_path=r"paid",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def paid(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        try:
            if order.pending_amount > 0:
                with transaction.atomic():
                    order_payments = Payment.objects.select_related(
                        "payment_method", "order__client"
                    ).filter(order_id=order.id, status=Payment.PaymentStatus.Pending)

                    for payment in order_payments:
                        payment_service = PaymentFactory.create_payment_service(
                            payment.payment_method.code_name
                        )
                        if (
                                payment_service.check_payment_status(payment, user)
                                != "completed"
                        ):
                            raise PaymentNotCompletedException()

                    if order.pending_amount == 0:
                        NotificationService.send_notification(
                            "Pagos pendientes pagados",
                            f"Los pagos de la solicitud <b>{order.id}</b> han sido <b>actualizados</b> correctamente",
                            [user, order.client],
                            "Informativo",
                            ["IN_APP", "WHATSAPP"],
                        )

                        next_status = models.OrderStatus.objects.get(code_name="completed")
                        handler = get_state_handler(next_status, order.current_status.status)
                        handler.handle_transition(order, None)

            return Response(
                serializers.OrderSerializer(order, context=self.get_serializer_context()).data,
                status=status.HTTP_200_OK,
            )
        except models.OrderStatus.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"], url_path=r"add-payments")
    def add_payments(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        with transaction.atomic():
            serializer = serializers.OrderAddPaymentSerializer(data=request.data)
            if serializer.is_valid(raise_exception=True):
                try:
                    amount = serializer.validated_data.get("amount")
                    payment_method = serializer.validated_data.get("payment_method_id")
                    paid = serializer.validated_data.get("paid")

                    amount /= payment_method.currency.exchange_rate

                    payment_service = PaymentFactory.create_payment_service(
                        payment_method.code_name
                    )
                    _, payment = payment_service.create_payment(
                        payment_method,
                        order,
                        Decimal(amount),
                    )
                    if paid:
                        payment_service.check_payment_status(payment, user)

                    NotificationService.send_notification(
                        "Nuevo pago realizado",
                        f"Se ha adicionado un nuevo pago al pedido <b>{order}</b>",
                        [user, order.client],
                        "Informativo",
                        ["IN_APP", "WHATSAPP"],
                    )

                    return Response(
                        serializers.OrderSerializer(order, context=self.get_serializer_context()).data,
                        status=status.HTTP_200_OK,
                    )
                except Payment.DoesNotExist:
                    return Response(status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"], url_path=r"completed")
    def completed_payment(self, request, pk=None):
        order = self.get_object()
        user = self.request.user
        try:
            with transaction.atomic():
                order_payment = Payment.objects.select_related("payment_method", "order__client").get(order_id=order.id)

                if order_payment.status == "pending":
                    payment_service = PaymentFactory.create_payment_service(order_payment.payment_method.code_name)
                    payment_service.complete_payment(order_payment)

                    NotificationService.send_notification(
                        "Solicitud pagada",
                        f"La solicitud de compra <b>{order.id}</b> ha sido <b>pagada</b> correctamente",
                        [user, order.client],
                        "Informativo",
                        ["IN_APP", "WHATSAPP"],
                    )
                return Response(
                    serializers.OrderSerializer(order, context=self.get_serializer_context()).data,
                    status=status.HTTP_200_OK,
                )

        except Payment.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        except models.OrderStatus.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"], url_path=r"cancel")
    def cancel_payment(self, request, pk=None):
        order = self.get_object()
        try:
            with transaction.atomic():
                order_payment = Payment.objects.select_related("payment_method", "order__client").get(order_id=order.id)
                payment_service = PaymentFactory.create_payment_service(order_payment.payment_method.code_name)
                payment_service.cancel_payment(order_payment)
                current_status = order.current_status.status
                cancelled_status = models.OrderStatus.objects.get(code_name="cancelled")
                handler = get_state_handler(cancelled_status, current_status)
                handler.handle_transition(order, "Pedido cancelado antes del pago")
                return Response(
                    serializers.OrderSerializer(
                        order, context={"request": request}
                    ).data,
                    status=status.HTTP_200_OK,
                )

        except Payment.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        except models.OrderStatus.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=True,
        methods=["post"],
        url_path=r"refund",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def refund(self, request, pk=None):
        raise NotImplemented
        order = self.get_object()
        user = self.request.user

        try:
            with transaction.atomic():
                order_payment = Payment.objects.select_related(
                    "currency", "order__client"
                ).get(order_id=order.id)

                if order_payment.status != Payment.PaymentStatus.Completed:
                    return Response(status=status.HTTP_400_BAD_REQUEST)

                order_payment.status = Payment.PaymentStatus.Refunded
                order_payment.save(update_fields=["status"])

                TransactionLog.objects.create(
                    transaction_id=order_payment.transaction_id,
                    payment_status=order_payment.status,
                    charge_for=user,
                    description=f"Pago reembolsado de la solicitud {order.id} por {order_payment.amount}. Reembolso realizado por {user.first_name} {user.last_name}",
                )
                order_shipping = OrderShipping.objects.get(order=order)
                shipping_amount = order_shipping.shipping_price
                refund_amount = order_payment.amount - shipping_amount
                wallet, created = Wallet.objects.select_for_update().get_or_create(
                    user_id=order.client_id, defaults={"amount": refund_amount}
                )

                if not created:
                    previous_amount = wallet.amount
                    wallet.amount += refund_amount
                    wallet.save(update_fields=["amount"])

                    WalletOperationalLog.objects.create(
                        transaction_id=f"REEMBOLSO_{order_payment.currency.initials}_{refund_amount}",
                        description=f"Reembolso de la solicitud {order.id}. Realizado por {user.first_name} {user.last_name}",
                        amount=refund_amount,
                        previous_amount=previous_amount,
                        exchange_rate=order_payment.exchange_rate,
                        exchange_rate_date=order_payment.exchange_rate_date,
                        wallet=wallet,
                        currency=order_payment.currency,
                        charge_for=user,
                    )

                cancelled_status = models.OrderStatus.objects.get(code_name="cancelled")
                handler = get_state_handler(cancelled_status, order.current_status.status)
                handler.handle_transition(order, "Pedido cancelado al realizarse un reembolso")

                order_products = models.OrderProducts.objects.select_related(
                    "product"
                ).filter(order_id=order.id)

                for op in order_products:
                    op.product.quantity += op.quantity
                    if op.product.quantity > 0:
                        op.product.active = True
                    op.product.save()

                NotificationService.send_notification(
                    "Solicitud reembolsada",
                    f"La solicitud de compra {order.id} ha sido reembolsada y cancelada. Monto reembolsado: <b>{refund_amount}</b>",
                    [order.client, user],
                    "Informativo",
                    ["IN_APP", "WHATSAPP"],
                )

                return Response(
                    serializers.OrderSerializer(
                        order, context=self.get_serializer_context()
                    ).data,
                    status=status.HTTP_200_OK,
                )

        except Payment.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        except OrderShipping.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

    @action(
        detail=True,
        methods=["patch"],
        url_path=r"change-status",
        permission_classes=[CustomPermissionFactory(["core.change_order_status"])],
    )
    def change_status(self, request, pk=None):
        with transaction.atomic():
            order = self.get_object()
            next_status = request.data.get("next_status")
            observations = request.data.get("observations")

            next_status = models.OrderStatus.objects.get(id=next_status)
            handler = get_state_handler(next_status, order.current_status.status)
            handler.handle_transition(order, observations)
            return Response(
                serializers.OrderSerializer(order, context=self.get_serializer_context()).data,
                status=status.HTTP_200_OK,
            )

    @action(
        detail=True,
        methods=["post"],
        url_path=r"clone",
        permission_classes=[ClientPermission],
    )
    def clone_order(self, request, pk=None):
        with transaction.atomic():
            order = self.get_object()
            client = order.client

            models.Cart.objects.filter(client_id=client.id).delete()

            order_products = models.OrderProducts.objects.filter(
                order_id=order.id
            ).select_related("product")

            cart_products_to_create = [
                models.Cart(
                    client_id=client.id,
                    quantity=op.quantity,
                    product=op.product,
                )
                for op in order_products
            ]
            if cart_products_to_create:
                models.Cart.objects.bulk_create(cart_products_to_create)
                total_amount = sum(
                    op.quantity * op.product.sell_price(client.fee)
                    for op in order_products
                )
                total_gross_weight = sum(
                    op.quantity * op.product.gross_weight for op in order_products
                )
                total_net_weight = sum(
                    op.quantity * op.product.net_weight for op in order_products
                )
            else:
                total_amount = total_gross_weight = total_net_weight = 0
            NotificationService.send_notification(
                "Solicitud clonada",
                f"La solicitud de compra {order.id} ha sido <b>clonada</b> y se encuentra disponible en su carrito de compra, para una próxima compra.",
                [client],
                "Informativo",
                ["IN_APP", "WHATSAPP"],
            )

            return Response(
                {
                    "total_amount": total_amount,
                    "total_gross_weight": total_gross_weight,
                    "total_net_weight": total_net_weight,
                    "products": serializers.CartSerializer(
                        cart_products_to_create if cart_products_to_create else [],
                        many=True,
                        context={"request": request},
                    ).data,
                },
                status=status.HTTP_200_OK,
            )

    @action(
        detail=True,
        methods=["get"],
        url_path=r"possible-to-mix",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def possible_to_mix(self, request, pk=None):
        order = self.get_object()
        orders_to_mix = (
            models.Order.objects.filter(client_id=order.client_id, merge=None)
            .exclude(id=order.id)
            .filter(
                Exists(
                    models.OrderTracking.objects.filter(
                        order_id=OuterRef("pk"),
                        status__initial_status=False,
                        status__final_status=False,
                    )
                )
            )
            .distinct()
        )

        return Response(
            serializers.OrderMinimalSerializer(
                orders_to_mix, many=True, context={"request": request}
            ).data,
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path=r"completed-and-paid",
        permission_classes=[CustomPermissionFactory(["core.manage_order"])],
    )
    def completed_and_paid(self, request, pk=None):

        latest_tracking_subquery = (
            models.OrderTracking.objects.filter(order=OuterRef("pk"))
            .order_by("-order_tracking_date")
            .values("id")[:1]
        )

        completed_and_paid_orders = (
            models.Order.objects.filter(
                order_trackings__id=Subquery(latest_tracking_subquery),
                order_trackings__status__code_name="completed",
                payment__status="completed",
            )
            .select_related("payment", "client")
            .prefetch_related("order_trackings__status")
        )

        return Response(
            serializers.OrderMinimalSerializer(
                completed_and_paid_orders, many=True, context={"request": request}
            ).data,
            status=status.HTTP_200_OK,
        )


class OrderProductsViewSet(ProtectedResourceViewSet):
    """
    Order Product model\n
    GET: Shows all Order Products created.\n
    POST: Adds a new Order Product.\n
    GET{id}: Retrieves a specific Order Product determined by id.\n
    PUT{id}: Modifies all fields of a specific Order Product determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Order Product determined by id.\n
    DELETE{id}: Deletes a specific Order Product determined by id.\n
    """

    permission_classes = [CustomPermissionFactory(["core.manage_orderproducts"])]
    queryset = models.OrderProducts.objects.all()
    serializer_class = serializers.OrderProductsSerializer

    def update(self, request, *args, **kwargs):
        with transaction.atomic():
            partial = kwargs.pop("partial", False)
            instance = self.get_object()
            order = instance.order
            if order.merge_id is not None:
                raise OrderUpdateException()
            order_statuses = models.OrderTracking.objects.filter(order_id=order.id)
            current_status = None
            if order_statuses.exists():
                current_status = order_statuses.first().status
            if current_status is not None:
                if current_status.final_status:
                    raise OrderUpdateException()
                new_quantity = (
                    request.data["quantity"] if "quantity" in request.data else None
                )
                if new_quantity is not None:
                    diference = instance.quantity - new_quantity
                    product = models.Product.objects.get(id=instance.product.id)
                    if product.quantity >= abs(diference):
                        product.quantity += diference
                        if product.quantity == 0:
                            product.active = False
                        product.save()
                        if (
                                product.has_wholesale_price
                                and new_quantity >= product.wholesale_minimum
                        ):
                            instance.price = product.wholesale_price
                        serializer = self.get_serializer(
                            instance, data=request.data, partial=partial
                        )
                        if serializer.is_valid(raise_exception=True):
                            self.perform_update(serializer)
                            NotificationService.send_notification(
                                "Solicitud actualizada",
                                f"Al producto <b>{product.name}</b> de la solicitud de compra <b>{order.id}</b> le ha sido actualizada la cantidad solicitada [<b>nueva cantidad: {new_quantity}</b>]",
                                [order.client],
                                "Informativo",
                                ["IN_APP", "WHATSAPP"],
                            )
                            return Response(
                                serializers.OrderSerializer(
                                    order, context={"request": request}
                                ).data,
                                status=status.HTTP_200_OK,
                            )
                    else:
                        raise OrderUpdateException()
            raise StatusUnAvalaibleException()

    def destroy(self, request, pk):
        with transaction.atomic():
            instance = self.get_object()
            product = instance.product
            order = instance.order
            self.perform_destroy(instance)
            NotificationService.send_notification(
                "Solicitud actualizada",
                f"A la solicitud de compra <b>{order.id}</b> se le ha eliminado el producto <b>{product.name}</b>",
                [order.client],
                "Informativo",
                ["IN_APP", "WHATSAPP"],
            )
            return Response(
                serializers.OrderSerializer(order, context={"request": request}).data,
                status=status.HTTP_200_OK,
            )


class OrderProfitReportViewSet(GenericAPIView):
    serializer_class = serializers.OrderProfitReportSerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date = serializer.validated_data.get("start_date")
        end_date = serializer.validated_data.get("end_date")

        if start_date > end_date:
            raise StartDateCanNotBeAfterEnddateException

        filters = {"order__creation_date__range": [start_date, end_date + datetime.timedelta(days=1)]}
        latest_status = Subquery(
            models.OrderTracking.objects.filter(order=OuterRef('order_id')).order_by('-id') \
                .values('status__code_name')[:1]
        )

        order_products = (
            models.OrderProducts.objects.filter(**filters)
            .select_related("order")
            .annotate(status=latest_status)
            .annotate(
                sale_amount=ExpressionWrapper(
                    F("quantity") * F("price"),
                    output_field=DecimalField(max_digits=15, decimal_places=2),
                ),
                sale_cost=ExpressionWrapper(
                    F("quantity") * F("cost"),
                    output_field=DecimalField(max_digits=15, decimal_places=2),
                ),
                absolute_margin=ExpressionWrapper(
                    F("quantity") * F("price") - F("quantity") * F("cost"),
                    output_field=DecimalField(max_digits=15, decimal_places=4),
                ),
                margin_percentual=ExpressionWrapper(
                    (F("quantity") * F("price") - F("quantity") * F("cost")) / (F("quantity") * F("price")),
                    output_field=DecimalField(max_digits=10, decimal_places=4),
                ),
            ).exclude(status__in=['cancelled', 'returned'])
            .order_by("-order__creation_date", "-order__id")
        )

        result = {}
        total_amount = 0
        total_cost = 0
        # total_profit = 0
        total_absolute_margin = 0

        for orderproduct in order_products:
            product_id = orderproduct.product_id
            data = result.get(
                product_id,
                {
                    "product": orderproduct.product,
                    "quantity": 0,
                    "sell_price": 0,
                    "sale_amount": 0,
                    "total_cost": 0,
                    "absolute_margin": 0,
                    "sales": [],
                },
            )

            p_quantity = orderproduct.quantity
            p_sell_price = orderproduct.price
            p_sale_amount = orderproduct.sale_amount
            p_total_cost = orderproduct.sale_cost
            p_absolute_margin = orderproduct.absolute_margin
            p_margin_percentual = orderproduct.margin_percentual
            p_profit_per_unit = orderproduct.price - orderproduct.cost

            data["sales"].append(
                {
                    "order_id": orderproduct.order.id,
                    "creation_date": orderproduct.order.creation_date,
                    "client_name": orderproduct.order.client.get_full_name,
                    "quantity": p_quantity,
                    "sell_price": p_sell_price,
                    "sale_amount": p_sale_amount,
                    "total_cost": p_total_cost,
                    "absolute_margin": p_absolute_margin,
                    "margin_percentual": p_margin_percentual,
                    "profit_per_unit": p_profit_per_unit,
                }
            )

            # total_profit += p_profit_per_unit * p_quantity
            total_amount += p_sale_amount
            total_cost += p_total_cost
            total_absolute_margin += p_absolute_margin

            data["quantity"] += p_quantity
            data["sell_price"] += p_sell_price
            data["sale_amount"] += p_sale_amount
            data["total_cost"] += p_total_cost
            data["absolute_margin"] += p_absolute_margin
            result[product_id] = data

        return Response(
            {
                "products": serializers.OrderProductProfit(result.values(), many=True).data,
                # "total_profit": total_profit,
                "total_amount": total_amount,
                "total_cost": total_cost,
                "total_absolute_margin": total_absolute_margin,
            },
            status=status.HTTP_200_OK,
        )


class BatchSalesReportView(GenericAPIView):
    def get(self, request):
        batch_id = request.query_params.get("batch_id")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        product_id = request.query_params.get("product_id")

        batch_items = models.BatchItem.objects.select_related(
            "batch", "product"
        ).prefetch_related("order_product__order_product__order")

        if batch_id:
            batch_items = batch_items.filter(batch_id=batch_id)

        if product_id:
            batch_items = batch_items.filter(product_id=product_id)

        if start_date and end_date:
            if start_date > end_date:
                raise StartDateCanNotBeAfterEnddateException
            batch_items = batch_items.filter(
                batch__received_date__gte=start_date, batch__received_date__lte=end_date
            )

        # if end_date:
        #     batch_items = batch_items.filter(batch__received_date__lte=end_date)

        report_data = []

        for batch_item in batch_items:
            product_batches = models.ProductBatch.objects.filter(
                batch_item=batch_item
            ).select_related("order_product__order")
            orders_info = product_batches.aggregate(
                total_orders=Count("order_product__order", distinct=True),
                total_sold_quantity=Coalesce(Sum("quantity"), Decimal("0.00")),
                total_revenue=Coalesce(
                    Sum(F("order_product__quantity") * F("order_product__price")),
                    Decimal("0.00"),
                ),
                avg_selling_price=Coalesce(
                    Avg("order_product__price"), Decimal("0.00")
                ),
            )
            total_cost = batch_item.cost_price * orders_info["total_sold_quantity"]
            total_profit = orders_info["total_revenue"] - total_cost
            profit_margin = Decimal("0.00")
            if orders_info["total_revenue"] > 0:
                profit_margin = (total_profit / orders_info["total_revenue"]) * 100
            remaining_quantity = batch_item.quantity - batch_item.quantity_sold

            report_data.append(
                {
                    "batch_id": batch_item.batch.id,
                    "batch_identificator": batch_item.batch.batch_identificator,
                    "received_date": batch_item.batch.received_date,
                    "exchange_rate": batch_item.batch.exchange_rate,
                    "invoice_number": batch_item.batch.invoice_number,
                    "part_code": batch_item.part_code,
                    "product_name": batch_item.name,
                    "product_id": batch_item.product.id if batch_item.product else None,
                    "cost_price": batch_item.cost_price,
                    "sale_price": batch_item.sale_price,
                    "total_quantity": batch_item.quantity,
                    "sold_quantity": orders_info["total_sold_quantity"],
                    "remaining_quantity": remaining_quantity,
                    "total_cost": total_cost,
                    "total_revenue": orders_info["total_revenue"],
                    "total_profit": total_profit,
                    "profit_margin_percentage": profit_margin,
                    "total_orders": orders_info["total_orders"],
                    "average_selling_price": orders_info["avg_selling_price"],
                }
            )
        report_data.sort(key=lambda x: (x["batch_identificator"], x["part_code"]))

        serializer = serializers.BatchSalesReportSerializer(report_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ReviewViewSet(ProtectedResourceViewSet, MultiplePermissionsView):
    """
    Review model\n
    GET: Shows all Reviews created.\n
    POST: Adds a new Review.\n
    GET{id}: Retrieves a specific Review determined by id.\n
    PUT{id}: Modifies all fields of a specific Review determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Review determined by id.\n
    DELETE{id}: Deletes a specific Review determined by id.\n
    """

    get_permission_classes = [AllowAny]
    post_permission_classes = [ClientPermission]
    queryset = models.Review.objects.all()
    serializer_class = serializers.ReviewSerializer
    search_fields = ["comment"]

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "review", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "review",
                "list",
                request.user,
                timeout=60 * 60 * 12,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "review", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "review",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("review")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("review")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("review")
        response = super().perform_destroy(instance)
        return response


class ProvinceViewSet(ProtectedResourceViewSet):
    """
    Province model\n
    GET: Shows all Provinces created.\n
    POST: Adds a new Province.\n
    GET{id}: Retrieves a specific Province determined by id.\n
    PUT{id}: Modifies all fields of a specific Province determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Province determined by id.\n
    DELETE{id}: Deletes a specific Province determined by id.\n
    """

    permission_classes = [
        IsAuthenticatedOrReadOnly | CustomPermissionFactory(["core.show_province"])
    ]
    queryset = models.Province.objects.all()
    serializer_class = serializers.ProvinceSerializer
    search_fields = ["name"]

    def list(self, request, *args, **kwargs):
        page = request.query_params.get("page")
        page_size = request.query_params.get("page_size")
        search_term = request.query_params.get("search", "")

        if search_term and search_term.strip():
            return super().list(request, *args, **kwargs)

        cache_kwargs = {"page": page, "page_size": page_size, "search": search_term}

        cached_data = NomenclatorCacheManager.get_cached_data(
            "province", "list", request.user, **cache_kwargs
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "province",
                "list",
                request.user,
                timeout=60 * 60 * 24 * 7,
                **cache_kwargs,
            )

        return response

    def retrieve(self, request, *args, **kwargs):
        cached_data = NomenclatorCacheManager.get_cached_data(
            "province", "retrieve", request.user, kwargs.get("pk")
        )

        if cached_data is not None:
            return Response(cached_data)

        response = super().retrieve(request, *args, **kwargs)

        if response.status_code == 200:
            NomenclatorCacheManager.set_cached_data(
                response.data,
                "province",
                "retrieve",
                request.user,
                pk=kwargs.get("pk"),
                timeout=60 * 60 * 24 * 30,
            )

        return response

    def perform_create(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("province")
        response = super().perform_create(serializer)
        return response

    def perform_update(self, serializer):
        NomenclatorCacheManager.invalidate_model_cache("province")
        response = super().perform_update(serializer)
        return response

    def perform_destroy(self, instance):
        NomenclatorCacheManager.invalidate_model_cache("province")
        response = super().perform_destroy(instance)
        return response


class MunicipalityViewSet(ProtectedResourceViewSet):
    """
    Municipality model\n
    GET: Shows all Municipality created.\n
    POST: Adds a new Municipality.\n
    GET{id}: Retrieves a specific Municipality determined by id.\n
    PUT{id}: Modifies all fields of a specific Municipality determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Municipality determined by id.\n
    DELETE{id}: Deletes a specific Municipality determined by id.\n
    """

    permission_classes = [
        IsAuthenticatedOrReadOnly | CustomPermissionFactory(["core.show_municipality"])
    ]
    queryset = models.Municipality.objects.all()
    serializer_class = serializers.MunicipalitySerializer
    filterset_class = filters.MunicipalityFilter
    search_fields = ["name"]


class ContactAddressViewSet(ProtectedResourceViewSet):
    """
    Contact Address model\n
    GET: Shows all Contact's addresses created.\n
    POST: Adds a new Contact's address.\n
    GET{id}: Retrieves a specific Contact's address determined by id.\n
    PUT{id}: Modifies all fields of a specific Contact's address determined by id.\n
    PATCH{id}: Partially modifies the fields of a specific Contact's address determined by id.\n
    DELETE{id}: Deletes a specific Contact's address determined by id.\n
    """

    permission_classes = [
        IsAuthenticatedOrReadOnly | ClientPermission | StaffPermission
    ]
    queryset = models.ContactAddress.objects.all()
    serializer_class = serializers.ContactAddressSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_staff:
            return queryset
        return queryset.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        """
        Create a new contact address
        """
        try:
            with transaction.atomic():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                contact_address = serializer.save()

                response_serializer = self.get_serializer(contact_address)
                return Response(
                    response_serializer.data, status=status.HTTP_201_CREATED
                )

        except Exception:
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(
        detail=True,
        methods=["get"],
        url_path=r"shipping-methods",
        permission_classes=[IsAuthenticatedOrReadOnly | ClientPermission],
    )
    def address_shipping_methods_rates(self, request, pk=None):
        with transaction.atomic():
            contact_address = self.get_object()
            municipality_id = contact_address.municipality.id
            shipping_rates = (
                ShippingRate.objects.filter(
                    active=True,
                    shipping_zone__active=True,
                    shipping_zone__municipalities__id=municipality_id,
                    shipping_method__active=True,
                )
                .select_related("shipping_zone", "shipping_method")
                .prefetch_related("shipping_zone__municipalities")
            )
            return Response(
                ShippingRateSerializer(
                    shipping_rates, many=True, context={"request": request}
                ).data,
                status=status.HTTP_200_OK,
            )


class ConfigAPIView(RetrieveUpdateAPIView):
    permission_classes = [
        ReadOnlyPermission | CustomPermissionFactory(["core.manage_config"])
    ]
    serializer_class = serializers.ConfigSerializer

    CACHE_KEY = "config_response"

    def get_object(self):
        return get_object_or_404(models.Config, pk=1)

    def get(self, request, *args, **kwargs):
        cached_response = cache.get(self.CACHE_KEY)
        if cached_response:
            return Response(cached_response)

        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data

        cache.set(self.CACHE_KEY, data, timeout=None)
        return Response(data)

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)

        cache.delete(self.CACHE_KEY)

        return response

    def put(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)
